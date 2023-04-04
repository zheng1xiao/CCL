#!/usr/bin/env python
# -*- coding: utf-8 -*-
# 导入模块 openpyxl
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
import copy

# 读取excel文件
# 括号中的字符串为你要比较的两个excel的路径，注意用“/”
wb_a = openpyxl.load_workbook('D:/lenovo/Documents/科研论文/co-teaching/CCL/hcommon_new.xlsx')
wb_b = openpyxl.load_workbook('D:/lenovo/Documents/科研论文/co-teaching/CCL/hcommon_old.xlsx')


# 定义一个方法来获取表格中某一列的内容，返回一个列表
# 将每一列输出为一个列表（temp表示列的名字）
def getIP(wb, temp):
    sheet = wb.active
    ip = []
    for cellobj in sheet[temp]:
        ip.append(cellobj.value)
    return ip


# 想比较哪几列就输入那几列的名称
list1 = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'N']
list2 = []  # 用于存每列不同的值
differ1 = {}  # 第一个文件中每列不同的列表组成字典
differ2 = {}  # 第二个文件中每列不同的列表组成字典
for temp1 in list1:
    # 获得ip列表
    ip_a = getIP(wb_a, temp1)
    ip_b = getIP(wb_b, temp1)
    # 将两个列表转换成集合
    aa = set(ip_a)
    bb = set(ip_b)
    # 找出两个列表的不同行，并转换成列表
    difference = list(aa ^ bb)
    # 打印出列表中的元素
    # 到这一步，两个表格中不同的数据已经被找出来了
    for i in difference:
        print (i)

    # 将不同行高亮显示
    # print ("开始第一张表" + "----" *10)
    del list2[0:]
    a = wb_a.active[temp1]
    for cellobj in a:
        if cellobj.value in difference:
            # print (cellobj.value)
            cellobj.font = Font(color=colors.BLACK, italic=True, bold=True)
            cellobj.fill = PatternFill("solid", fgColor="DDDDDD")
            list2.append(cellobj.value)
    if list2 != []:
        differ1[temp1] = copy.deepcopy(list2)
    # print ("开始第二张表" + "----" *10)
    del list2[0:]
    b = wb_b.active[temp1]
    for cellobj in b:
        if cellobj.value in difference:
            # print (cellobj.value)
            cellobj.font = Font(color=colors.BLACK, italic=True, bold=True)
            cellobj.fill = PatternFill("solid", fgColor="DDDDDD")
            list2.append(cellobj.value)
    if list2 != []:
        differ2[temp1] = copy.deepcopy(list2)

print(differ1.items())
print(differ2.items())
wb_a.save('D:/lenovo/Documents/科研论文/co-teaching/CCL/hcommon_new.xlsx')
wb_b.save('D:/lenovo/Documents/科研论文/co-teaching/CCL/hcommon_old.xlsx')
